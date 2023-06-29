import openpyxl
import os
from win32com.client import Dispatch
import time


class OperationExcel:

    def __init__(self, file_path=None, file_name=None, sheetname=None):
        # 当前路径
        current_path = os.path.abspath(os.path.dirname(__file__))
        # print(current_path)
        # 上一级路径（父级路径）
        parent_path = os.path.dirname(current_path)
        if file_path is None:
            file_path = parent_path + "\\upload_file\\"
        if file_name is None:
            file_name = "团险-正常标准团单被保人清单"
        self.excel_file = file_path + file_name + '.xlsx'
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(self.excel_file)
        self.new_file = None
        # self.sheets = self.wb.sheetnames  # 获取表格所有的sheet 名称
        if sheetname is None:
            self.sheet = self.wb.worksheets[0]
        else:
            self.sheet = self.wb[sheetname]
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column

    # 获取单元格的值
    def get_cell_value(self, row, column):
        row, column = int(row), int(column)
        if row > self.max_row:
            return 'error, row is out of limit'
        if column > self.max_column:
            return 'error, column is out of limit'
        return self.sheet.cell(row, column).value

    # 保存单元格值
    def save_to_excel_by_cell(self, row, column, value):

        row, column = int(row), int(column)
        self.sheet.cell(row, column).value = value
        self.wb.save(self.excel_file)

    # 获取某一列的值
    def get_column_value(self, column_name):
        column = self.sheet[column_name]
        column_data = {}

        for i in range(1, self.max_row):
            key = str(column[i].value)
            column_data[key] = i + 1

        return column_data

    # 打开和关闭，重新保存文件
    def resave_excel(self, file_name):
        xlApp = Dispatch('Excel.Application')
        xlApp.Visible = False  # 打开wps时是否在桌面展示
        workbook = xlApp.Workbooks.Open(self.excel_file)  # 打开的路径
        self.new_file = self.file_path + file_name + '.xlsx'
        workbook.SaveAs(self.new_file)  # 保存的路径
        xlApp.Quit()

    # 删除文件
    def delete_excel_file(self, delete_file):
        time.sleep(3)
        os.remove(delete_file)

if __name__ == '__main__':
    test = OperationExcel()
    print(test.max_row)

    # test.save_to_excel_by_cell(2, 2, 'test1')
    # test.save_to_excel_by_cell(3, 3, 'test2')
    # print(test.get_cell_value(6, 6))
    print(test.get_cell_value(6, 1))
